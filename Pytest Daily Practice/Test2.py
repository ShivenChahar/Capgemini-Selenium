# from selenium.webdriver import Chrome, ChromeOptions
# from selenium.webdriver.common.by import By
#
#
# @pytest.mark.parametrize('username, password',[('standard_user','secret_sauce'),('error_user','secret_sauce')])
# def test_login(username, password):
#     options = ChromeOptions()
#     options.add_experimental_option('detach', True)
#     driver = Chrome(options=options)
#     driver.get("https://www.saucedemo.com")
#     driver.maximize_window()
#     driver.implicitly_wait(10)
#     driver.find_element(By.ID, "user-name").send_keys(username)
#     driver.find_element(By.ID, "password").send_keys(password)
#     driver.find_element(By.ID, "login-button").click()
#     assert "inventory" in driver.current_url, "Login failed"
#     driver.quit()

from openpyxl.workbook import Workbook

import openpyxl

wb = openpyxl.Workbook()
sheetName = "sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else :
    ws = wb.create_sheet(sheetName)

ws['A1'] = 'user'
ws['B1'] = 'password'

# wb.save('sample.xlsx')   # It will save inn the local directory

ws.append(['user1', '123'])
ws.append(['user2', '456'])
ws.append(['user3', '789'])
ws.append(['user4', '789'])
ws.append(['user5', '789'])

wb.save('sample.xlsx')  # To save the file

for row in ws.iter_rows(values_only=True):    # To iterate in the Excel Sheet
    print(row)
